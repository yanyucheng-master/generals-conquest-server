# -*- coding: utf-8 -*-
"""
读取 card_data_export 下最新的 .xlsx，并输出模板样式要点：
- 工作表列表
- 冻结窗格、筛选范围
- 表头行样式（加粗、填充色、对齐、自动换行）
- 列宽（前若干列）
"""

from __future__ import annotations

from pathlib import Path
import os
import json

import openpyxl
from openpyxl.utils import get_column_letter


def main() -> None:
    root = Path(__file__).resolve().parents[1]
    export_dir = root / "card_data_export"
    if not export_dir.exists():
        raise SystemExit(f"card_data_export not found: {export_dir}")

    files = list(export_dir.glob("*.xlsx"))
    if not files:
        raise SystemExit(f"no xlsx in {export_dir}")

    files.sort(key=lambda p: p.stat().st_mtime, reverse=True)
    path = files[0]
    print("use_xlsx:", str(path))

    wb = openpyxl.load_workbook(path)
    print("sheets_json:", json.dumps(wb.sheetnames, ensure_ascii=True))

    def inspect_sheet(ws) -> dict:
        # column widths
        widths = {}
        for col, dim in ws.column_dimensions.items():
            if dim.width:
                widths[col] = dim.width

        # find first non-empty row within top 30 as header candidate
        header_row = None
        for r in range(1, 31):
            if any(ws.cell(r, c).value is not None for c in range(1, ws.max_column + 1)):
                header_row = r
                break

        headers = []
        if header_row is not None:
            for c in range(1, min(ws.max_column, 30) + 1):
                cell = ws.cell(header_row, c)
                val = cell.value
                if val is None:
                    continue
                fill = None
                try:
                    fill = cell.fill.fgColor.value
                except Exception:
                    fill = None
                headers.append(
                    {
                        "col": get_column_letter(c),
                        "value": str(val),
                        "bold": bool(cell.font.bold),
                        "font_color": getattr(cell.font.color, "value", None),
                        "fill": fill,
                        "align_h": cell.alignment.horizontal,
                        "align_v": cell.alignment.vertical,
                        "wrap": bool(cell.alignment.wrap_text),
                    }
                )

        return {
            "title": ws.title,
            "freeze_panes": str(ws.freeze_panes) if ws.freeze_panes else None,
            "auto_filter": ws.auto_filter.ref if getattr(ws, "auto_filter", None) else None,
            "col_widths_first_30": dict(list(widths.items())[:30]),
            "header_row_guess": header_row,
            "header_cells": headers,
            "row2_values_first_12": [ws.cell(2, c).value for c in range(1, 13)] if ws.max_row >= 2 else None,
        }

    summary = {name: inspect_sheet(wb[name]) for name in wb.sheetnames}
    print("template_json:", json.dumps(summary, ensure_ascii=True, indent=2))


if __name__ == "__main__":
    main()

