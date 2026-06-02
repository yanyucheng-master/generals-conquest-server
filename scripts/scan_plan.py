# -*- coding: utf-8 -*-
import openpyxl
from pathlib import Path

wb = openpyxl.load_workbook(Path(r'c:\Users\YYC\Desktop\01_将领征服_现行卡牌数据表.xlsx'), data_only=True)
out = Path(__file__).resolve().parent.parent / 'card_data_export'
out.mkdir(exist_ok=True)
lines = []
for i, ws in enumerate(wb.worksheets):
    lines.append(f'Sheet {i}: {ws.title}')
    for row in ws.iter_rows(values_only=True):
        if row and any(x and '计划' in str(x) for x in row):
            lines.append(f'  FOUND: {row}')
    # last rows of 通用
    if '通用' in ws.title or i == 4:
        lines.append('  --- all rows ---')
        for row in ws.iter_rows(values_only=True):
            if any(row):
                lines.append(f'  {row}')
(out / 'scan_plan.txt').write_text('\n'.join(lines), encoding='utf-8')
print('done', len(lines))
