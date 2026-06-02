# -*- coding: utf-8 -*-
"""法术类型卡的子类型统一为「法术卡」，并刷新汇总页。"""
import json
import openpyxl
from openpyxl.styles import Font, Alignment
from pathlib import Path
from datetime import datetime

OUT_DIR = Path(__file__).resolve().parent.parent / 'card_data_export'
SRC_DESKTOP = Path(r'c:\Users\YYC\Desktop\01_将领征服_现行卡牌数据表.xlsx')

# 优先更新已汇总文件，否则桌面源文件
candidates = list(OUT_DIR.glob('将领征服_卡牌数据表_已汇总_*.xlsx'))
XLSX_PATH = candidates[-1] if candidates else SRC_DESKTOP

HEADER = ['ID', '名称', '费用', '品质', '类型', '子类型', '攻击', '生命', '护甲', '描述', '技能', '阵营']
TYPE_COL = 5   # 1-based: 类型
SUBTYPE_COL = 6  # 子类型
HEADER_ROW = 2
DATA_START = 3
SPELL_TYPE = '法术'
SPELL_SUBTYPE = '法术卡'


def update_sheet(ws, stats):
    for row in range(DATA_START, ws.max_row + 1):
        card_type = ws.cell(row=row, column=TYPE_COL).value
        if card_type and str(card_type).strip() == SPELL_TYPE:
            old = ws.cell(row=row, column=SUBTYPE_COL).value
            ws.cell(row=row, column=SUBTYPE_COL, value=SPELL_SUBTYPE)
            name = ws.cell(row=row, column=2).value
            cid = ws.cell(row=row, column=1).value
            stats['updated'].append({
                'id': cid, 'name': name, 'old_subtype': old, 'new_subtype': SPELL_SUBTYPE,
                'sheet': ws.title,
            })


def rebuild_summary(wb, stats):
    ws_sum = wb.worksheets[-1]
    all_rows = []
    for ws in wb.worksheets[1:-1]:
        for row in range(DATA_START, ws.max_row + 1):
            cid = ws.cell(row=row, column=1).value
            if cid is None:
                continue
            try:
                int(cid)
            except (TypeError, ValueError):
                continue
            row_vals = [ws.cell(row=row, column=c).value for c in range(1, 13)]
            all_rows.append(row_vals)
    all_rows.sort(key=lambda r: int(r[0]))

    ws_sum.cell(row=1, column=1, value=f'全部卡牌汇总 ({len(all_rows)}张)')
    for col, h in enumerate(HEADER, 1):
        cell = ws_sum.cell(row=HEADER_ROW, column=col, value=h)
        cell.font = Font(bold=True)
    if ws_sum.max_row > DATA_START - 1:
        ws_sum.delete_rows(DATA_START, max(0, ws_sum.max_row - (DATA_START - 1)))
    for i, row_vals in enumerate(all_rows, start=DATA_START):
        for col, val in enumerate(row_vals, 1):
            ws_sum.cell(row=i, column=col, value=val)
    stats['summary_count'] = len(all_rows)


def main():
    wb = openpyxl.load_workbook(XLSX_PATH)
    stats = {'updated': [], 'file': str(XLSX_PATH)}

    for i, ws in enumerate(wb.worksheets):
        if i == 0 or i == len(wb.worksheets) - 1:
            continue
        update_sheet(ws, stats)

    rebuild_summary(wb, stats)

    out_name = f'将领征服_卡牌数据表_已汇总_{datetime.now().strftime("%Y%m%d")}.xlsx'
    out_path = OUT_DIR / out_name
    wb.save(out_path)

    src_ok = False
    if SRC_DESKTOP.exists():
        try:
            wb.save(SRC_DESKTOP)
            src_ok = True
        except PermissionError:
            pass

    stats['output'] = str(out_path)
    stats['source_updated'] = src_ok
    stats['spell_count'] = len(stats['updated'])

    (OUT_DIR / 'sync_report.json').write_text(
        json.dumps(stats, ensure_ascii=False, indent=2), encoding='utf-8'
    )

    lines = [
        f'已更新法术子类型为「{SPELL_SUBTYPE}」共 {len(stats["updated"])} 张',
        f'输出: {out_path.name}',
        '',
        'ID | 名称 | 原子类型',
        '---|---|---',
    ]
    for u in stats['updated']:
        lines.append(f"{u['id']} | {u['name']} | {u['old_subtype']}")
    (OUT_DIR / '法术子类型更新清单.txt').write_text('\n'.join(lines), encoding='utf-8')

    readme = f"""将领：征服 — 卡牌数据表（已汇总 v2）
更新时间：{datetime.now().strftime('%Y-%m-%d %H:%M')}

【本版变更】
1. 各阵营分表 → 全卡牌汇总 已同步
2. 新卡「计划」ID=82 | 2费铜法术 | drawCard | 通用
3. 所有「类型=法术」的卡，子类型统一为「法术卡」（共 {len(stats['updated'])} 张）

【主文件】
{out_path.name}

【待你确认后】
再同步进 cards.ts 与游戏逻辑
"""
    (OUT_DIR / '说明.txt').write_text(readme, encoding='utf-8')
    print(json.dumps({
        'spell_updated': len(stats['updated']),
        'total': stats['summary_count'],
        'output': str(out_path),
    }, ensure_ascii=False))


if __name__ == '__main__':
    main()
