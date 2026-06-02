# -*- coding: utf-8 -*-
"""将各阵营分表同步到全卡牌汇总，规范化新卡「计划」，输出 Excel。"""
import json
import openpyxl
from openpyxl.styles import Font, Alignment
from pathlib import Path
from datetime import datetime
from copy import copy

OUT_DIR = Path(__file__).resolve().parent.parent / 'card_data_export'
OUT_DIR.mkdir(exist_ok=True)
DEFAULT_SRC = Path(r'c:\Users\YYC\Desktop\01_将领征服_现行卡牌数据表.xlsx')

def pick_source_xlsx() -> Path:
    """
    优先使用 card_data_export 下最新的 .xlsx 作为模板（保留用户美化格式）。
    若不存在则回退到桌面默认源文件。
    """
    cands = sorted(OUT_DIR.glob('*.xlsx'), key=lambda p: p.stat().st_mtime, reverse=True)
    return cands[0] if cands else DEFAULT_SRC

HEADER = ['ID', '名称', '费用', '品质', '类型', '子类型', '攻击', '生命', '护甲', '描述', '技能', '阵营']
# 工作簿顺序：封面, 帝国军团, 荒野游侠, 奥术学院, 通用, 全卡牌汇总
FACTION_WS_INDEX = [1, 2, 3, 4]
NEW_CARD_ID = 82  # 计划


def normalize_skills(skills_raw: str) -> str:
    s = str(skills_raw or '').strip()
    if s in ('-', 'None', '', 'nan'):
        return '-'
    return s


def row_to_card(row):
    """从一行解析卡牌；首列非数字时视为新增草稿。"""
    if not row or row[1] is None:
        return None
    name = str(row[1]).strip()
    if not name or name == '名称':
        return None
    try:
        cid = int(row[0])
    except (TypeError, ValueError):
        return {
            'id': NEW_CARD_ID if name == '计划' else None,
            'name': name,
            'cost': int(row[2] or 0),
            'quality': str(row[3] or '铜').strip(),
            'type': str(row[4] or '法术').strip(),
            'subtype': str(row[5] or '魔法').strip() if row[5] else '魔法',
            'atk': int(row[6] or 0),
            'hp': int(row[7] or 0),
            'armor': int(row[8] or 0),
            'desc': str(row[9] or '').strip() or '抽取1',
            'skills': 'drawCard' if name == '计划' else normalize_skills(row[10]),
            'faction': str(row[11] or '通用').strip() if row[11] else '通用',
            'is_new': True,
        }
    subtype = str(row[5] or '').strip()
    if subtype in ('None', 'nan'):
        subtype = ''
    return {
        'id': cid,
        'name': name,
        'cost': int(row[2]),
        'quality': str(row[3]).strip(),
        'type': str(row[4]).strip(),
        'subtype': subtype,
        'atk': int(row[6] or 0),
        'hp': int(row[7] or 0),
        'armor': int(row[8] or 0),
        'desc': str(row[9] or '').strip(),
        'skills': normalize_skills(row[10]),
        'faction': str(row[11] or '').strip(),
        'is_new': False,
    }


def parse_faction_sheet(ws):
    cards = []
    for row in ws.iter_rows(min_row=3, values_only=True):
        c = row_to_card(row)
        if c and c.get('id') is not None:
            cards.append(c)
    return cards


def card_to_row(c):
    return [
        c['id'], c['name'], c['cost'], c['quality'], c['type'], c['subtype'],
        c['atk'], c['hp'], c['armor'], c['desc'], c['skills'], c['faction'],
    ]


def write_summary_sheet(ws, all_cards):
    # 标题
    ws.cell(row=1, column=1, value=f'全部卡牌汇总 ({len(all_cards)}张)')
    for col, h in enumerate(HEADER, 1):
        cell = ws.cell(row=2, column=col, value=h)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
    if ws.max_row > 2:
        ws.delete_rows(3, ws.max_row - 2)
    for i, c in enumerate(all_cards, start=3):
        for col, val in enumerate(card_to_row(c), 1):
            ws.cell(row=i, column=col, value=val)


def write_plan_to_faction_sheet(ws):
    """在通用表写入规范化的「计划」行（ID 82）。"""
    plan_row = None
    plan_row_idx = None
    for idx, row in enumerate(ws.iter_rows(min_row=3, values_only=True), start=3):
        if row and row[1] and str(row[1]).strip() == '计划':
            plan_row_idx = idx
            break
    plan = {
        'id': NEW_CARD_ID,
        'name': '计划',
        'cost': 2,
        'quality': '铜',
        'type': '法术',
        'subtype': '魔法',
        'atk': 0,
        'hp': 0,
        'armor': 0,
        'desc': '抽取1',
        'skills': 'drawCard',
        'faction': '通用',
    }
    if plan_row_idx:
        for col, val in enumerate(card_to_row(plan), 1):
            ws.cell(row=plan_row_idx, column=col, value=val)
    else:
        r = ws.max_row + 1
        for col, val in enumerate(card_to_row(plan), 1):
            ws.cell(row=r, column=col, value=val)
    # 更新表头计数
    if ws.cell(row=1, column=1).value and '通用' in str(ws.cell(row=1, column=1).value):
        ws.cell(row=1, column=1, value='通用 (4张)')


def update_cover(ws, total: int):
    for r in range(1, 20):
        for c in range(1, 6):
            v = ws.cell(row=r, column=c).value
            if v and isinstance(v, str):
                if '总计' in v or '总卡牌' in v:
                    ws.cell(
                        row=r, column=c,
                        value=f'总计{total}张卡牌 | 三阵营 + 通用 | {datetime.now().strftime("%Y年%m月")}（已同步）',
                    )
                if '帝国军团' in v and '张' in v:
                    pass  # 保持分阵营说明由分表标题维护


def main():
    src_path = pick_source_xlsx()
    wb = openpyxl.load_workbook(src_path)

    all_cards = []
    faction_counts = {}
    ws_names = [wb.worksheets[i].title for i in FACTION_WS_INDEX]

    for i, idx in enumerate(FACTION_WS_INDEX):
        ws = wb.worksheets[idx]
        cards = parse_faction_sheet(ws)
        # 通用表：追加「计划」
        if idx == 4:
            write_plan_to_faction_sheet(ws)
            cards = parse_faction_sheet(ws)
        faction_counts[ws.title] = len(cards)
        all_cards.extend(cards)

    all_cards.sort(key=lambda c: c['id'])
    plan = next((c for c in all_cards if c['name'] == '计划'), None)

    # 更新汇总表（最后一个 sheet）
    ws_sum = wb.worksheets[-1]
    write_summary_sheet(ws_sum, all_cards)
    update_cover(wb.worksheets[0], len(all_cards))

    out_name = f'将领征服_卡牌数据表_已汇总_{datetime.now().strftime("%Y%m%d")}.xlsx'
    out_path = OUT_DIR / out_name
    wb.save(out_path)
    # 按约定：card_data_export 目录仅保留 .xlsx，因此不输出 json/txt 报告
    print(json.dumps(report, ensure_ascii=False, indent=2))


if __name__ == '__main__':
    main()
