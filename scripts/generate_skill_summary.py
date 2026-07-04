# -*- coding: utf-8 -*-
"""
从代码与卡牌数据生成「现行技能汇总表」。
输出至 skill_summary_export/（xlsx/md/json）；含技能表的合并卡牌表仍输出至 card_data_export/
"""
from __future__ import annotations

import json
import re
from collections import defaultdict
from datetime import date, datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent.parent
CARDS_TS = ROOT / "src" / "data" / "cards.ts"
GAME_TS = ROOT / "src" / "types" / "game.ts"
ENGINE = ROOT / "src" / "engine" / "gameEngine.ts"
OUT_DIR_XLSX = ROOT / "card_data_export"          # 只放 .xlsx（按约定）
OUT_DIR_OTHER = ROOT / "skill_summary_export"     # md/json/说明放这里
TODAY = datetime.now().strftime("%Y%m%d_%H%M")

# 实现状态：已实现 | 部分实现 | 未实现 | 仅DIY/无官方卡
SKILL_REGISTRY: dict[str, dict] = {
    "flashStrike": {
        "name": "闪击",
        "status": "已实现",
        "trigger": "部署时",
        "logic": "部署后立即 pickAutoTarget 并 applyDamage 一次；标记 flashStrikeUsed。联机同步 flashStrike 结果。",
        "params": "无等级参数",
        "files": "gameEngine.ts deployUnit / mirrorEnemyDeploy",
    },
    "magicSwap": {
        "name": "魔术",
        "status": "已实现",
        "trigger": "部署时 / 施放法术",
        "logic": "随机选敌方场上两格（非总部）交换单位；疾风步可指定目标与另一随机单位交换。",
        "params": "无",
        "files": "gameEngine.ts deployUnit, castSpell",
    },
    "pursuit": {
        "name": "追击",
        "status": "未实现",
        "trigger": "设计：敌方位移或前线部署时",
        "logic": "仅在 getSkillLevel 保留解析规则；无触发代码。现行官方卡未引用。",
        "params": "desc 中「追击N点」",
        "files": "—",
    },
    "nimble": {
        "name": "灵动",
        "status": "未实现",
        "trigger": "设计：位置变化时",
        "logic": "仅 getSkillLevel 解析；无位移监听。现行官方卡未引用。",
        "params": "desc「灵动…加N」",
        "files": "—",
    },
    "bleed": {
        "name": "流血",
        "status": "已实现",
        "trigger": "攻击命中 / 回合开始结算",
        "logic": "攻击时严格读取描述「流血N」并令 target.bleed += N（未写数值时默认1）；敌方回合开始 resolveBleed 扣除当前层数伤害并将层数减半。可作用于总部。",
        "params": "desc「流血N」",
        "files": "gameEngine.ts applyDamage, resolveBleed, startTurn",
    },
    "tear": {
        "name": "撕裂",
        "status": "已实现",
        "trigger": "攻击命中 / 荒野呼唤法术",
        "logic": "攻击前若目标已有流血层数，先造成等于流血层数的真实伤害，再附加新流血。荒野呼唤对单目标执行撕裂+毒爆。",
        "params": "无",
        "files": "gameEngine.ts applyDamage, castSpell",
    },
    "poison": {
        "name": "中毒",
        "status": "已实现",
        "trigger": "攻击命中",
        "logic": "攻击时严格读取描述「中毒N」并令 target.poison += N（未写数值时默认2）。中毒单位的强运、战术指挥、射击指挥失效。",
        "params": "desc「中毒N」",
        "files": "gameEngine.ts applyDamage, getAttackBonus, hasLuckyOnBoard",
    },
    "poisonBurst": {
        "name": "毒爆",
        "status": "已实现",
        "trigger": "攻击命中 / 荒野呼唤",
        "logic": "攻击时若目标已有中毒，造成等于中毒层数的伤害并清零；荒野呼唤同理。",
        "params": "无",
        "files": "gameEngine.ts applyDamage, castSpell",
    },
    "growth": {
        "name": "生长",
        "status": "已实现",
        "trigger": "受到治疗时",
        "logic": "治疗生效后 atk/maxHp/hp 各 +getSkillLevel(growth)。",
        "params": "desc「生长…加N」",
        "files": "gameEngine.ts startTurn(heal), castSpell(heal/aoeHeal)",
    },
    "balance": {
        "name": "均衡",
        "status": "已实现",
        "trigger": "战斗阶段开始 / 施放均衡法术",
        "logic": "将己方手牌中所有士兵卡费用改为 getSkillLevelFromCard(balance)（默认4，军备扩充等为3）。新抽士兵受场上均衡单位影响。",
        "params": "desc「均衡N」",
        "files": "gameEngine.ts combatPhaseStart, castSpell, drawCardRaw",
    },
    "pierce": {
        "name": "贯穿",
        "status": "已实现",
        "trigger": "造成伤害时",
        "logic": "物理伤害可击破护甲破碎保护，溢出伤害正常扣生命。",
        "params": "无",
        "files": "gameEngine.ts applyDamage",
    },
    "piercePlus": {
        "name": "强化贯穿",
        "status": "已实现",
        "trigger": "造成伤害时",
        "logic": "物理伤害无视目标护甲数值。",
        "params": "无",
        "files": "gameEngine.ts applyDamage",
    },
    "physResist": {
        "name": "物抗",
        "status": "已实现",
        "trigger": "受到物理伤害时",
        "logic": "伤害 -= getSkillLevel(physResist)；强击可无视。",
        "params": "desc「物抗N」",
        "files": "gameEngine.ts applyDamage",
    },
    "magicResist": {
        "name": "法抗",
        "status": "已实现",
        "trigger": "受到魔法伤害时",
        "logic": "法术/魔法攻击减伤；法力贯穿/强击可无视。",
        "params": "desc「法抗N」",
        "files": "gameEngine.ts applyDamage, castSpell",
    },
    "allResist": {
        "name": "全抗",
        "status": "已实现",
        "trigger": "受到伤害时",
        "logic": "物理/魔法均减伤；强击、法力贯穿（法术）可无视。",
        "params": "desc「全抗N」",
        "files": "gameEngine.ts applyDamage, castSpell",
    },
    "tacticCmd": {
        "name": "战术指挥",
        "status": "已实现",
        "trigger": "攻击时计算加成",
        "logic": "己方场上带战术指挥的单位使所有近战友军攻击 +N（含自己）。",
        "params": "desc「战术指挥N」",
        "files": "gameEngine.ts getUnitAttackBonus, GameBoard",
    },
    "shootCmd": {
        "name": "射击指挥",
        "status": "已实现",
        "trigger": "攻击时计算加成",
        "logic": "己方弓箭友军攻击 +N。",
        "params": "desc「射击指挥N」",
        "files": "gameEngine.ts getUnitAttackBonus, GameBoard",
    },
    "lucky": {
        "name": "强运",
        "status": "已实现",
        "trigger": "随机伤害判定",
        "logic": "己方场上有未中毒、未沉默的强运单位时，己方随机伤害取区间最大值。",
        "params": "无",
        "files": "gameEngine.ts rollDamage",
    },
    "spellReflect": {
        "name": "法术反弹",
        "status": "部分实现",
        "trigger": "受到法术伤害时",
        "logic": "castSpell 中若目标有 spellReflect 且法术无法力贯穿：记录反弹日志，部分分支仍可能扣血。单位普攻 magicBullet 走 manaPierce 判定（与 magicPierce 不同步）。",
        "params": "无",
        "files": "gameEngine.ts castSpell, applyDamage",
    },
    "magicBoost": {
        "name": "法力增幅",
        "status": "已实现",
        "trigger": "部署时 / 命运编织者",
        "logic": "部署时己方所有 subtype=魔法 的单位攻击 +N（含自己），buff 在敌方回合开始移除；一次性 magicBoostUsed。refreshBoost 可重置并再触发。",
        "params": "desc「增幅N」「法力增幅N」",
        "files": "gameEngine.ts deployUnit, castSpell(refreshBoost)",
    },
    "focusFire": {
        "name": "集火",
        "status": "已实现",
        "trigger": "施放集火令 / 攻击阶段",
        "logic": "设置 focusTarget；攻击阶段优先攻击该目标；集火令对目标额外 1 点伤害（HQ 或单位）。",
        "params": "无",
        "files": "gameEngine.ts castSpell, runAttackPhase",
    },
    "spear": {
        "name": "长矛",
        "status": "已实现",
        "trigger": "射程计算",
        "logic": "近战射程变为 1 或 2。",
        "params": "无",
        "files": "gameEngine.ts getUnitRange",
    },
    "shortBow": {
        "name": "短弓",
        "status": "已实现",
        "trigger": "射程与选目标",
        "logic": "弓箭手可攻击距离 1，且可打 HQ/任意距离（与普通弓箭仅距离2不同）。",
        "params": "无",
        "files": "gameEngine.ts getUnitRange, getTargetsInRange",
    },
    "conceal": {
        "name": "隐蔽",
        "status": "部分实现",
        "trigger": "被选为攻击目标时",
        "logic": "底线带 conceal 的单位不会被近战/弓箭/狙击选为目标（引擎有过滤）。现行官方卡池无此技能。",
        "params": "无",
        "files": "gameEngine.ts getTargetsInRange",
    },
    "immune": {
        "name": "免疫",
        "status": "已实现",
        "trigger": "法术选目标 / 消灭",
        "logic": "带免疫的单位不能被法术指定（含末日审判）；useGame/AI 选目标时过滤。",
        "params": "无",
        "files": "gameEngine.ts castSpell, useGame.ts",
    },
    "counter": {
        "name": "反击",
        "status": "已实现",
        "trigger": "被近战物理命中后",
        "logic": "未沉默时，对攻击者反击一次（非反击链）。",
        "params": "无",
        "files": "gameEngine.ts applyDamage",
    },
    "intimidate": {
        "name": "叱吓",
        "status": "已实现",
        "trigger": "攻击加成计算",
        "logic": "敌方近战/弓箭单位攻击 -getSkillLevel(intimidate)。",
        "params": "desc「叱吓N」",
        "files": "gameEngine.ts getAttackBonus",
    },
    "strongStrike": {
        "name": "强击",
        "status": "已实现",
        "trigger": "造成伤害时",
        "logic": "无视物抗/法抗/全抗。",
        "params": "无",
        "files": "gameEngine.ts applyDamage",
    },
    "bounty": {
        "name": "悬赏",
        "status": "已实现",
        "trigger": "单位被击杀",
        "logic": "击杀方 +1 金币。",
        "params": "无",
        "files": "gameEngine.ts applyDamage",
    },
    "fly": {
        "name": "飞翔",
        "status": "已实现",
        "trigger": "受到物理攻击时",
        "logic": "50% 闪避近战，25% 闪避弓箭（精准/防空可破）。射程不限。",
        "params": "无",
        "files": "gameEngine.ts applyDamage, getUnitRange",
    },
    "antiAir": {
        "name": "防空",
        "status": "已实现",
        "trigger": "攻击飞翔单位",
        "logic": "必中且伤害 +getSkillLevel(antiAir)。",
        "params": "desc「防空N」",
        "files": "gameEngine.ts applyDamage",
    },
    "riddleRealm": {
        "name": "谜境",
        "status": "已实现",
        "trigger": "部署时",
        "logic": "若手牌有法术，设置 riddleActive=true，表示谜语法术已激活；圣光可清除对应法术。",
        "params": "无",
        "files": "gameEngine.ts applyRiddleRealm",
    },
    "dive": {
        "name": "俯冲",
        "status": "已实现",
        "trigger": "物理攻击时",
        "logic": "攻击者同时有 fly 时伤害 +getSkillLevel(dive)。",
        "params": "desc「俯冲N」",
        "files": "gameEngine.ts applyDamage",
    },
    "silence": {
        "name": "沉默",
        "status": "已实现",
        "trigger": "施放带 silence 的法术 / 部署带 silence 的单位",
        "logic": "法术：敌方全体 silenceTurns = getSkillLevelFromCard(silence)。士兵部署（如禁咒法师）：按 desc 沉默等级作用于敌方全场。混乱风暴另设 spellOnlyNextTurn 禁部署。",
        "params": "desc「沉默N」",
        "files": "gameEngine.ts castSpell, applyDeployUnitPassives",
    },
    "agile": {
        "name": "疾行",
        "status": "已实现",
        "trigger": "部署阶段（无选手牌时）",
        "logic": "点击带疾行且未使用的己方单位，再点击相邻空格移动一次；回合开始重置 agileUsed。",
        "params": "无",
        "files": "gameEngine.ts moveAgileUnit, useGame.ts",
    },
    "fog": {
        "name": "迷雾",
        "status": "部分实现",
        "trigger": "施放法术",
        "logic": "仅日志提示隐藏信息，无实际遮蔽 UI/信息。现行官方卡无。",
        "params": "无",
        "files": "gameEngine.ts castSpell",
    },
    "stealth": {
        "name": "隐踪",
        "status": "未实现",
        "trigger": "部署（未开放）",
        "logic": "未开放 / 未实现 / 暂不参与正式平衡。代码中仅设置 isStealthed 标记，法术选目标未完整校验隐踪。",
        "params": "无",
        "files": "gameEngine.ts cardToUnit（部分标记）",
    },
    "ambush": {
        "name": "伏击",
        "status": "已实现",
        "trigger": "受到攻击前",
        "logic": "对攻击者造成 getSkillLevel(ambush) 伤害；若击杀则取消本次攻击。",
        "params": "desc「伏击N」",
        "files": "gameEngine.ts applyDamage",
    },
    "precision": {
        "name": "精准",
        "status": "已实现",
        "trigger": "狙击攻击",
        "logic": "无视闪避/飞翔；伤害 +getSkillLevel(precision)。",
        "params": "desc「精准N」",
        "files": "gameEngine.ts applyDamage",
    },
    "disguise": {
        "name": "伪装",
        "status": "已实现",
        "trigger": "被选为目标",
        "logic": "狙击/随机子类型无法选中带伪装单位。",
        "params": "无",
        "files": "gameEngine.ts applyDamage",
    },
    "holyLight": {
        "name": "圣光",
        "status": "已实现",
        "trigger": "部署带圣光的单位",
        "logic": "若敌方 riddleActive，弃掉其手牌中全部法术；否则随机弃 1 张法术。",
        "params": "无",
        "files": "gameEngine.ts applyHolyLight, applyDeployUnitPassives",
    },
    "interest": {
        "name": "利息",
        "status": "已实现",
        "trigger": "回合开始",
        "logic": "每个带 interest 的单位按 getSkillLevel(interest) 从 desc「利息N」加金币。",
        "params": "desc「利息N」",
        "files": "gameEngine.ts startTurn",
    },
    "revenge": {
        "name": "复仇",
        "status": "已实现",
        "trigger": "友方单位阵亡",
        "logic": "其他友方带复仇单位立即 pickAutoTarget 并攻击一次。",
        "params": "无",
        "files": "gameEngine.ts applyDamage",
    },
    "dodge": {
        "name": "闪避",
        "status": "已实现",
        "trigger": "受到物理伤害",
        "logic": "50% 完全闪避（精准除外）。",
        "params": "无",
        "files": "gameEngine.ts applyDamage",
    },
    "extract": {
        "name": "萃取",
        "status": "已实现",
        "trigger": "击杀敌方单位",
        "logic": "攻击方金币 +getSkillLevel(extract)。",
        "params": "desc「萃取N」",
        "files": "gameEngine.ts applyDamage",
    },
    "taunt": {
        "name": "嘲讽",
        "status": "已实现",
        "trigger": "自动选目标",
        "logic": "有嘲讽单位时只从嘲讽单位中选目标。",
        "params": "无",
        "files": "gameEngine.ts pickAutoTarget",
    },
    "jamming": {
        "name": "干扰",
        "status": "已实现",
        "trigger": "狙击阶段",
        "logic": "敌方场上有干扰时，玩家狙击目标随机化。",
        "params": "无",
        "files": "gameEngine.ts resolveSniper, GameBoard",
    },
    "manaPierce": {
        "name": "法力贯穿（内部兼容别名）",
        "status": "已实现",
        "trigger": "与 magicPierce 相同",
        "logic": "旧数据兼容 ID；applyDamage / castSpell 中与 magicPierce 等价处理，不作为独立官方技能展示。",
        "params": "无",
        "files": "gameEngine.ts applyDamage, castSpell",
    },
    "magicPierce": {
        "name": "法力贯穿",
        "status": "已实现",
        "trigger": "法术伤害 / 单位魔法攻击",
        "logic": "castSpell 与 applyDamage 均识别 magicPierce（及别名 manaPierce），可跳过法抗/法术反弹。",
        "params": "无",
        "files": "gameEngine.ts castSpell, applyDamage",
    },
    "magicBullet": {
        "name": "魔力子弹",
        "status": "已实现",
        "trigger": "狙击攻击",
        "logic": "狙击攻击伤害类型改为魔法。",
        "params": "无",
        "files": "gameEngine.ts applyDamage",
    },
    "lifesteal": {
        "name": "吸血",
        "status": "已实现",
        "trigger": "造成伤害后",
        "logic": "攻击方 HQ 恢复等于实际伤害的生命。",
        "params": "无",
        "files": "gameEngine.ts applyDamage",
    },
    "magicDmg": {
        "name": "法术伤害（内部效果码）",
        "status": "已实现",
        "trigger": "施放法术",
        "logic": "仅用于引擎识别法术伤害分支，不属于官方技能、不会作为技能特效展示。默认 2 伤；法力风暴 6；毒针 1（受伤+1）；天火降临固定 12 点打总部；混乱风暴对全场所有单位 1 伤（单独分支）。",
        "params": "部分按卡名硬编码",
        "files": "gameEngine.ts castSpell, damageAllFieldUnits",
    },
    "shield": {
        "name": "护盾术",
        "status": "已实现",
        "trigger": "施放法术",
        "logic": "目标 HQ 或友方单位 +2 护甲（固定值）。现行官方卡无。",
        "params": "固定 +2",
        "files": "gameEngine.ts castSpell",
    },
    "heal": {
        "name": "治疗",
        "status": "已实现",
        "trigger": "回合开始(单位) / 法术",
        "logic": "回合开始：战地医师等治疗血量最低友军；法术：对指定目标恢复 getSkillLevelFromCard。",
        "params": "desc「治疗N」",
        "files": "gameEngine.ts startTurn, castSpell",
    },
    "aoeHeal": {
        "name": "群体治疗",
        "status": "已实现",
        "trigger": "施放法术",
        "logic": "HQ 与全体友军恢复固定量。现行官方卡无。",
        "params": "默认 2",
        "files": "gameEngine.ts castSpell",
    },
    "armor": {
        "name": "护甲",
        "status": "已实现",
        "trigger": "部署时",
        "logic": "unit.armor = getSkillLevel(armor)；受物理伤先扣护甲。",
        "params": "desc「护甲N」",
        "files": "gameEngine.ts deployUnit",
    },
    "destroy": {
        "name": "消灭",
        "status": "已实现",
        "trigger": "末日审判等法术",
        "logic": "直接移除目标单位；免疫可挡；不能指定总部。",
        "params": "无",
        "files": "gameEngine.ts castSpell",
    },
    "refreshBoost": {
        "name": "刷新增幅",
        "status": "已实现",
        "trigger": "命运编织者",
        "logic": "己方场上未沉默且未用过的 magicBoost 单位再次对魔法友军 +攻。",
        "params": "无",
        "files": "gameEngine.ts castSpell",
    },
    "cleanseSilence": {
        "name": "净化沉默",
        "status": "已实现",
        "trigger": "命运编织者",
        "logic": "己方全体 silenceTurns 清零。",
        "params": "无",
        "files": "gameEngine.ts castSpell",
    },
    "drawCard": {
        "name": "抽牌",
        "status": "已实现",
        "trigger": "法术 / 部署",
        "logic": "计划抽1；军情急报3；急行军/重整2；知己知彼补至同手牌且至少2；其余按 desc「抽取N」。部署时解析 desc 抽牌。",
        "params": "卡名 + desc",
        "files": "gameEngine.ts executeDrawCardSpell, deployUnit",
    },
    "healHQ": {
        "name": "治疗总部",
        "status": "已实现",
        "trigger": "重整旗鼓",
        "logic": "从卡牌描述中的「总部恢复X」或「总部回复X」解析实际恢复量；未写数值时默认 3。现行重整旗鼓恢复 3。",
        "params": "desc「总部恢复X / 总部回复X」，默认 3",
        "files": "gameEngine.ts castSpell",
    },
    "discard": {
        "name": "弃牌",
        "status": "已实现",
        "trigger": "截获密信",
        "logic": "随机弃对手 1 张手牌；联机带 discardIdx 同步。",
        "params": "固定 1 张",
        "files": "gameEngine.ts castSpell",
    },
}


def parse_skills_from_cards() -> dict[str, list[dict]]:
    text = CARDS_TS.read_text(encoding="utf-8")
    usage: dict[str, list[dict]] = defaultdict(list)
    for m in re.finditer(
        r"\{\s*id:\s*(\d+),\s*name:\s*'([^']+)',\s*cost:\s*\d+,\s*quality:\s*'[^']+',\s*type:\s*'([^']+)',\s*subtype:\s*'[^']+',\s*atk:\s*\d+,\s*hp:\s*\d+,\s*armor:\s*\d+,\s*desc:\s*'([^']*)',\s*skills:\s*\[([^\]]*)\],\s*faction:\s*'([^']+)'",
        text,
    ):
        cid, name, ctype, desc, skills_raw, faction = m.groups()
        skills = [s.strip().strip("'") for s in skills_raw.split(",") if s.strip()]
        for sk in skills:
            usage[sk].append(
                {"id": int(cid), "name": name, "type": ctype, "faction": faction, "desc": desc}
            )
    return usage


def parse_skill_type_order() -> list[str]:
    text = GAME_TS.read_text(encoding="utf-8")
    order = []
    for m in re.finditer(r"\|\s*'(\w+)'\s*//", text):
        order.append(m.group(1))
    return order


def engine_has(skill: str) -> bool:
    eng = ENGINE.read_text(encoding="utf-8")
    return f"'{skill}'" in eng or f'"{skill}"' in eng


def write_markdown(path: Path, skill_order: list[str], usage: dict) -> None:
    lines = [
        "# 将领：征服 — 现行技能汇总",
        "",
        f"生成日期：{date.today().isoformat()}",
        "",
        "数据来源：`src/types/game.ts`、`src/engine/gameEngine.ts`、`src/data/cards.ts`。",
        "",
        "## 状态说明",
        "",
        "| 状态 | 含义 |",
        "|------|------|",
        "| 已实现 | 引擎有完整逻辑，与卡面基本一致 |",
        "| 部分实现 | 有代码但缺分支、数值不符或仅 UI/日志 |",
        "| 未实现 | 类型/卡面存在但无有效逻辑 |",
        "",
        "## 技能总表",
        "",
        "| 技能ID | 中文名 | 实现状态 | 触发时机 | 实际运行逻辑 | 参数来源 | 代码位置 | 官方卡引用数 |",
        "|--------|--------|----------|----------|--------------|----------|----------|--------------|",
    ]
    for sk in skill_order:
        meta = SKILL_REGISTRY.get(sk)
        if not meta:
            meta = {
                "name": sk,
                "status": "未知",
                "trigger": "—",
                "logic": "未在汇总脚本登记",
                "params": "—",
                "files": "—",
            }
        cards = usage.get(sk, [])
        lines.append(
            f"| `{sk}` | {meta['name']} | {meta['status']} | {meta['trigger']} | {meta['logic']} | {meta['params']} | {meta['files']} | {len(cards)} |"
        )

    lines.extend(["", "## 官方卡牌技能引用", ""])
    for sk in skill_order:
        cards = usage.get(sk, [])
        if not cards:
            continue
        lines.append(f"### {SKILL_REGISTRY.get(sk, {}).get('name', sk)} (`{sk}`)")
        lines.append("")
        for c in sorted(cards, key=lambda x: x["id"]):
            lines.append(f"- [{c['id']}] {c['name']}（{c['faction']}·{c['type']}）— {c['desc']}")
        lines.append("")

    lines.extend(
        [
            "## 备注",
            "",
            "混乱风暴使用 PlayerState.spellOnlyNextTurn 标记「下回合仅可打法术」；与 riddleActive 配合圣光/谜境。",
            "",
        ]
    )
    path.write_text("\n".join(lines), encoding="utf-8")


def write_xlsx(path: Path, skill_order: list[str], usage: dict) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "技能汇总"

    headers = [
        "技能ID",
        "中文名",
        "实现状态",
        "触发时机",
        "实际运行逻辑",
        "参数/数值来源",
        "主要代码位置",
        "引擎含引用",
        "官方卡数量",
        "引用卡牌(ID-名称)",
    ]
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True)
    for col, h in enumerate(headers, 1):
        cell = ws.cell(1, col, h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    status_colors = {
        "已实现": "C6EFCE",
        "部分实现": "FFEB9C",
        "未实现": "FFC7CE",
        "未知": "D9D9D9",
    }

    row = 2
    for sk in skill_order:
        meta = SKILL_REGISTRY.get(sk, {})
        cards = usage.get(sk, [])
        card_str = "；".join(f"{c['id']}-{c['name']}" for c in sorted(cards, key=lambda x: x["id"]))
        vals = [
            sk,
            meta.get("name", sk),
            meta.get("status", "未知"),
            meta.get("trigger", ""),
            meta.get("logic", ""),
            meta.get("params", ""),
            meta.get("files", ""),
            "是" if engine_has(sk) else "否",
            len(cards),
            card_str or "（官方卡池未使用）",
        ]
        for col, v in enumerate(vals, 1):
            c = ws.cell(row, col, v)
            c.alignment = Alignment(wrap_text=True, vertical="top")
            if col == 3:
                c.fill = PatternFill("solid", fgColor=status_colors.get(str(v), "FFFFFF"))
        row += 1

    widths = [14, 10, 10, 14, 48, 18, 28, 10, 10, 36]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"

    ws2 = wb.create_sheet("卡牌技能明细")
    h2 = ["卡牌ID", "名称", "阵营", "类型", "描述", "技能ID", "技能中文", "该技能实现状态"]
    for col, h in enumerate(h2, 1):
        cell = ws2.cell(1, col, h)
        cell.fill = header_fill
        cell.font = header_font
    r2 = 2
    text = CARDS_TS.read_text(encoding="utf-8")
    for m in re.finditer(
        r"\{\s*id:\s*(\d+),\s*name:\s*'([^']+)'.*?skills:\s*\[([^\]]*)\],\s*faction:\s*'([^']+)'",
        text,
        re.DOTALL,
    ):
        cid, name, skills_raw, faction = m.group(1), m.group(2), m.group(3), m.group(4)
        type_m = re.search(r"type:\s*'([^']+)'", m.group(0))
        desc_m = re.search(r"desc:\s*'([^']*)'", m.group(0))
        ctype = type_m.group(1) if type_m else ""
        desc = desc_m.group(1) if desc_m else ""
        skills = [s.strip().strip("'") for s in skills_raw.split(",") if s.strip()]
        if not skills:
            ws2.cell(r2, 1, int(cid))
            ws2.cell(r2, 2, name)
            ws2.cell(r2, 3, faction)
            ws2.cell(r2, 4, ctype)
            ws2.cell(r2, 5, desc)
            ws2.cell(r2, 6, "—")
            r2 += 1
            continue
        for sk in skills:
            meta = SKILL_REGISTRY.get(sk, {})
            for col, v in enumerate(
                [
                    int(cid),
                    name,
                    faction,
                    ctype,
                    desc,
                    sk,
                    meta.get("name", sk),
                    meta.get("status", "未知"),
                ],
                1,
            ):
                ws2.cell(r2, col, v)
            r2 += 1

    for i, w in enumerate([8, 14, 10, 8, 40, 14, 10, 12], 1):
        ws2.column_dimensions[get_column_letter(i)].width = w
    ws2.freeze_panes = "A2"

    wb.save(path)


def main() -> None:
    OUT_DIR_XLSX.mkdir(parents=True, exist_ok=True)
    OUT_DIR_OTHER.mkdir(parents=True, exist_ok=True)
    usage = parse_skills_from_cards()
    order = parse_skill_type_order()
    # 补全 registry 有但 type 未列出的
    for sk in SKILL_REGISTRY:
        if sk not in order:
            order.append(sk)

    md_path = OUT_DIR_OTHER / f"现行技能汇总_{TODAY}.md"
    xlsx_path = OUT_DIR_OTHER / f"现行技能汇总_{TODAY}.xlsx"
    json_path = OUT_DIR_OTHER / f"现行技能汇总_{TODAY}.json"

    write_markdown(md_path, order, usage)
    write_xlsx(xlsx_path, order, usage)

    export = {
        "generated": date.today().isoformat(),
        "skills": [
            {
                "id": sk,
                **SKILL_REGISTRY.get(sk, {}),
                "official_card_count": len(usage.get(sk, [])),
                "official_cards": usage.get(sk, []),
                "engine_reference": engine_has(sk),
            }
            for sk in order
        ],
    }
    json_path.write_text(json.dumps(export, ensure_ascii=False, indent=2), encoding="utf-8")

    # 若存在已汇总卡牌 xlsx，追加技能表
    summaries = sorted(OUT_DIR_XLSX.glob("将领征服_卡牌数据表_已汇总_*.xlsx"))
    if summaries:
        src = summaries[-1]
        wb = openpyxl.load_workbook(src)
        if "技能汇总" in wb.sheetnames:
            del wb["技能汇总"]
        ws = wb.create_sheet("技能汇总", 0)
        tmp = openpyxl.load_workbook(xlsx_path, read_only=True)
        src_ws = tmp["技能汇总"]
        for row in src_ws.iter_rows(values_only=True):
            ws.append(row)
        tmp.close()
        out_merged = OUT_DIR_XLSX / src.name.replace(".xlsx", f"_含技能表_{TODAY}.xlsx")
        wb.save(out_merged)
        print("merged", out_merged)

    # card_data_export 只保留 .xlsx，因此说明输出到 OUT_DIR_OTHER
    readme = OUT_DIR_OTHER / "说明.txt"
    extra = f"""
4. 现行技能汇总（{date.today().isoformat()}）
   - 现行技能汇总_{TODAY}.md / .xlsx / .json
   - 含实现状态（已实现/部分实现/未实现）与 gameEngine 实际逻辑说明
"""
    if readme.exists():
        txt = readme.read_text(encoding="utf-8")
        if "现行技能汇总" not in txt:
            readme.write_text(txt.rstrip() + extra, encoding="utf-8")
    else:
        readme.write_text(extra.strip(), encoding="utf-8")

    print("md", md_path)
    print("xlsx", xlsx_path)
    print("json", json_path)


if __name__ == "__main__":
    main()
